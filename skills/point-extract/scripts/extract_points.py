"""포인트 지급용 파일 추출 스크립트
사용법: python extract_points.py <입력파일.xlsx> [--output <출력파일.xlsx>]

예약자 정보 Excel에서 recipient_no, reservationNoWithCount, voucher1
3개 컬럼을 추출하고, 수량만큼 행을 확장하여 새 Excel로 저장한다.
"""

import argparse
import os
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl")
    sys.exit(1)


def normalize_phone(value):
    """전화번호 정규화: 숫자형 → 문자열, 하이픈 포맷 적용 (010-1234-5678)"""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        s = str(int(value))
    else:
        # 기존 하이픈/공백 제거 후 숫자만 추출
        s = "".join(c for c in str(value).strip() if c.isdigit())
    # 10자리이고 10으로 시작하면 0 추가 (예: 1012345678 → 01012345678)
    if len(s) == 10 and s.startswith("10"):
        s = "0" + s
    # 11자리 010 번호 → 010-XXXX-XXXX
    if len(s) == 11 and s.startswith("010"):
        s = f"{s[:3]}-{s[3:7]}-{s[7:]}"
    return s


def find_available_path(base_path):
    """동일 파일명이 있으면 _v2, _v3 ... suffix를 붙여 새 경로 반환 (덮어쓰기 방지)"""
    if not os.path.exists(base_path):
        return base_path
    name, ext = os.path.splitext(base_path)
    version = 2
    while True:
        candidate = f"{name}_v{version}{ext}"
        if not os.path.exists(candidate):
            return candidate
        version += 1


def extract_points(input_path, output_path=None):
    if not os.path.exists(input_path):
        print(f"❌ 파일을 찾을 수 없습니다: {input_path}")
        sys.exit(1)

    # 출력 경로 결정
    if output_path is None:
        dir_name = os.path.dirname(input_path) or "."
        today = datetime.now().strftime("%Y%m%d")
        output_path = os.path.join(dir_name, f"point_output_{today}.xlsx")
    output_path = find_available_path(output_path)

    wb_in = openpyxl.load_workbook(input_path)
    ws_in = wb_in.active

    # 헤더에서 필수 컬럼 인덱스 탐색
    headers = [cell.value for cell in ws_in[1]]
    required = {"예약번호": "reservation", "여행자 연락처": "phone", "수량": "qty"}
    col_map = {}
    for i, h in enumerate(headers):
        if h in required:
            col_map[required[h]] = i

    missing = set(required.values()) - col_map.keys()
    if missing:
        label_map = {v: k for k, v in required.items()}
        missing_labels = [label_map[m] for m in missing]
        print(f"❌ 필수 컬럼 누락: {', '.join(missing_labels)}")
        sys.exit(1)

    # 출력 워크북 생성
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.append(["recipient_no", "reservationNoWithCount", "voucher1"])

    total_rows = 0
    skipped = 0

    for row in ws_in.iter_rows(min_row=2, values_only=True):
        reservation_no = str(row[col_map["reservation"]] or "")
        phone = normalize_phone(row[col_map["phone"]])
        raw_qty = row[col_map["qty"]]

        # 수량 검증
        try:
            qty = int(raw_qty)
        except (TypeError, ValueError):
            qty = 0

        if qty <= 0:
            skipped += 1
            print(f"⚠️ 수량 0 또는 비정상 → 건너뜀: 예약번호={reservation_no}")
            continue

        for _ in range(qty):
            ws_out.append([phone, reservation_no, ""])
            total_rows += 1

    wb_out.save(output_path)

    print(f"✅ 완료: {total_rows}행 생성 → {output_path}")
    if skipped:
        print(f"⚠️ {skipped}건 건너뜀 (수량 0 또는 누락)")

    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="포인트 지급용 파일 추출")
    parser.add_argument("input", help="입력 Excel 파일 경로")
    parser.add_argument("--output", "-o", help="출력 Excel 파일 경로 (미지정 시 자동 생성)")
    args = parser.parse_args()
    extract_points(args.input, args.output)
